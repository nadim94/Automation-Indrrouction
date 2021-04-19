from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.options import Options
# OpenPyXl need to install for this
import openpyxl

path = r"C:\Users\Md. Nadim Kaysar\Downloads\New folder\test.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active  # workbook.get_sheet_names("sheet1")
rows = sheet.max_row
cols = sheet.max_column
print(rows)
print(cols)

for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(row=r, column=c).value, end="   ")
    print()
