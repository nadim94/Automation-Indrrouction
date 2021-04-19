from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.options import Options
# OpenPyXl need to install for this
import openpyxl


def getRowcount(file, sheet_Name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_Name)
    return (sheet.max_row)

def getColumncount(file, sheet_Name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_Name)
    return (sheet.max_column)


def readData(file, sheet_Name, rownum, colnumber):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_Name)
    return sheet.cell(row=rownum, column=colnumber).value


def writeData(file, sheet_Name, rownumber, colnumber, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_Name)
    sheet.cell(row=rownumber, column=colnumber).value = data
    workbook.save(file)

